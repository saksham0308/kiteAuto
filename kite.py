import datetime
import openpyxl
from openpyxl import load_workbook
import time
from kiteconnect import KiteTicker
from openpyxl.styles import PatternFill
from selenium import webdriver
from kiteconnect import KiteConnect
from selenium.webdriver.common.by import By
import pandas as pd
import json

my_file = "loginData.json"
with open(my_file, "r") as file:
    myData = json.load(file)
kite = KiteConnect(api_key=myData.get("api_key"))
# Redirect the user to the login url obtained
# from kite.login_url(), and receive the request_token
# from the registered redirect url after the login flow.
# Once you have the request_token, obtain the access_token
# as follows.

# This is not needed if chromedriver is already on your path:

options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
driver.get(kite.login_url())
driver.implicitly_wait(10)

driver.find_element(By.ID, 'userid').send_keys(myData.get("userid"))
driver.find_element(By.ID, "password").send_keys(myData.get("password"))
driver.find_element(By.CLASS_NAME, "button-orange").click()
time.sleep(20)

start = driver.current_url.find("request_token=") + 14
end = driver.current_url.find("&action")
if end == -1:
    myToken = driver.current_url[start:]
else:
    myToken = driver.current_url[start:end]
print(myToken)

data = kite.generate_session(myToken, api_secret=myData.get("api_secret"))
kite.set_access_token(data["access_token"])
print("Session started")

kws = KiteTicker(myData.get("api_key"), data["access_token"])
file_source = r'BNData.xlsx'
a = 3
used = False
curRow = 0
df = 0
df_transposed = 0
sheet_name = 0
last_price = 0
flag = False
strike_price = 0
option_strike = ""
option_strikeNext1 = ""
prev1Token = 0
prev2Token = 0
next1Token = 0
next2Token = 0
option_strikePrev2 = ""
format1 = "%d-%m-%Y"
format2 = '%H:%M:%S'


# Get instruments
# with open('readme.csv', 'w') as f:
#     f.write(str(kite.instruments('NFO')))
def round_to_nearest_hundred(number):
    remainder = number % 100

    # Determine whether to round up or down based on the remainder
    if remainder >= 50:
        rounded_number = number + (100 - remainder)
    else:
        rounded_number = number - remainder

    return rounded_number


def create_option_strike(symbol, date, rounded_number, option_type='PE'):
    return f"{symbol}{date}{rounded_number:05d}{option_type}"


contract_prices = {
}


def on_ticks(ws, ticks):
    global a, prev1Token, prev2Token, next1Token, next2Token
    global used
    global curRow
    global df
    global df_transposed
    global contract_prices
    global sheet_name
    global last_price
    global flag
    global strike_price
    global option_strike
    global option_strikeNext1
    global option_strikePrev2
    global format1
    global format2

    # if not flag:
    #     t = datetime.datetime.now()
    #     latestPrice = 0
    #     for tick in ticks:
    #         if str(tick['instrument_token']) == str(myData.get('BANKNIFTYTOKEN')):
    #             latestPrice = tick['last_price']
    #     print("time before flag false ", t.minute, t.second, t.microsecond / 1000)
    #     if latestPrice != (myData.get("bankNiftyLastPrice")):
    #         print("Latest Price")
    #         print(latestPrice)
    #         t = datetime.datetime.now()
    #         print("time after flag true ", t.minute, t.second, t.microsecond / 1000)
    #         symbol = myData.get("symbol")
    #         date = myData.get("date")
    #         strike_price = round_to_nearest_hundred(int(latestPrice))
    #         print(strike_price)
    #         option_strike = create_option_strike(symbol, date, strike_price)
    #         print(option_strike)
    #         flag = True
    #         print(flag)

    # Callback to receive ticks.
    print(ticks)
    now = datetime.datetime.now()
    print(now.second)

    contract_prices['Date and Time'].append(now.strftime(format2))
    for key, value in contract_prices.items():
        if str(key) == 'Date and Time':
            continue
        check = 1
        for tick in ticks:
            # print(str(key) + " "+str(tick['instrument_token']))
            if str(key) == str(tick['instrument_token']):
                jj = 10
                for ii in range(50):
                    jj = jj + 2
                    contract_prices[key].append(tick['last_price'] + jj)
                    jj = jj - 1
                    contract_prices[key].append(tick['last_price'] + jj)
                check = 0
                break
        if check == 1:
            contract_prices[key].append(0)


def extract_values(json_file):
    now = datetime.datetime.today()
    contract_prices[str("Date and Time")] = [now.strftime(format1)]
    with open(json_file, 'r') as file:
        data = json.load(file)
        values = list(data.values())
        keys = list(data.keys())
        for key, value in zip(keys, values):
            contract_prices[str(value)] = [str(key)]
    return values


json_filename = 'token.json'


def on_connect(ws, response):
    # Callback on successful connect.
    # Subscribe to a list of instrument_tokens (RELIANCE and ACC here).
    tokens = extract_values(json_filename)
    ws.subscribe(tokens)
    ws.set_mode(ws.MODE_FULL, tokens)
    # 260105==banknifty


# Assign the callbacks.
kws.on_ticks = on_ticks
kws.on_connect = on_connect

# Infinite loop on the main thread. Nothing after this will run.
# You have to use the pre-defined callbacks to manage subscriptions.
kws.connect(threaded=True)
# Place an order
# BuyOrder

while myData.get("OrderPlacementEnabled"):
    t = datetime.datetime.now()
    if t.minute >= int(myData.get("timeInMinutes")) and (t.microsecond / 1000) >= 50:
        latestPrice = int(kite.quote("NSE:NIFTY BANK")["NSE:NIFTY BANK"]['last_price'])
        symbol = myData.get("symbol")
        date = myData.get("date")
        strike_price = round_to_nearest_hundred(int(latestPrice))
        print(strike_price)
        option_strike = create_option_strike(symbol, date, strike_price)
        print(option_strike)
        print("Time before placing the order", t.minute, t.second, t.microsecond / 1000)
        try:
            order_id = kite.place_order(tradingsymbol=option_strike,
                                        exchange=kite.EXCHANGE_NFO,
                                        transaction_type=kite.TRANSACTION_TYPE_BUY,
                                        quantity=int(myData.get("quantity")),
                                        variety=kite.VARIETY_REGULAR,
                                        order_type=kite.ORDER_TYPE_MARKET,
                                        product=kite.PRODUCT_MIS,
                                        validity=kite.VALIDITY_DAY)

            t = datetime.datetime.now()
            print("Time after placing the order and getting the order id ", t.minute, t.second, t.microsecond / 1000)
            print("The order id is: {}".format(order_id))
        except Exception as e:
            print("Order placement failed: {}".format(e))
        print("Breaking buyOrder Loop")
        break

t = datetime.datetime.now()
print("Time before 10 second sleep", t.minute, t.second, t.microsecond / 1000)
time.sleep(9)
t = datetime.datetime.now()
print("Time after 10 second sleep", t.minute, t.second, t.microsecond / 1000)

# kite.cancel_order(variety=kite.VARIETY_AMO,
#                   order_id='231120000001027')


# SEll ORDER

try:
    orderBook = kite.orders()
    print(orderBook)
    print(orderBook[-1]['status'])
    t = datetime.datetime.now()
    print("Time before selling the order", t.minute, t.second, t.microsecond / 1000)
    if orderBook[-1]['status'] == 'COMPLETE' and myData.get("OrderPlacementEnabled"):
        orderid2 = kite.place_order(tradingsymbol=option_strike,
                                    exchange=kite.EXCHANGE_NFO,
                                    transaction_type=kite.TRANSACTION_TYPE_SELL,
                                    quantity=int(myData.get("quantity")),
                                    variety=kite.VARIETY_REGULAR,
                                    order_type=kite.ORDER_TYPE_MARKET,
                                    product=kite.PRODUCT_MIS,
                                    validity=kite.VALIDITY_DAY)
        t = datetime.datetime.now()
        print("Time after selling the order", t.minute, t.second, t.microsecond / 1000)
        print("The order id is: {}".format(orderid2))
except Exception as e:
    print(e)

print("DONE Buy/Sell")
print("Starting extra 5 second sleep")
time.sleep(5)
print("Session Ended")


def append_df_to_excel(df, excel_path):
    try:
        # Load the existing Excel workbook
        book = load_workbook(excel_path)
        currentMonth = str(t.today().strftime("%B"))
        # Get the existing sheet or create a new one
        sheet = book[currentMonth] if currentMonth in book.sheetnames else book.create_sheet(title=currentMonth)

        # Find the last used row dynamically
        last_used_row = sheet.max_row if sheet.max_row > 1 else 1

        for key in df.index:
            last_used_row += 1
            sheet.cell(row=last_used_row, column=1, value=key)

            # Write values from the DataFrame to subsequent columns
            for idx, value in enumerate(df.loc[key].astype(str).values, start=2):
                sheet.cell(row=last_used_row, column=idx, value=value)

        # Save the changes to the existing Excel file
        book.save(excel_path)
        print(f"DataFrame appended to {excel_path} under {sheet_name} starting from row {last_used_row}")

    except Exception as e:
        print(f"Error appending DataFrame to {excel_path}: {e}")


dynamic_range = max(len(values) if isinstance(values, list) else 0 for values in contract_prices.values())
df = pd.DataFrame.from_dict(contract_prices, orient='index')

# Display the DataFrame
print(df)
append_df_to_excel(df, 'BNData.xlsx')


# Load the existing workbook

def is_numeric(value):
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


workbook = openpyxl.load_workbook('BNData.xlsx')
worksheet = workbook.active
# Define fill colors
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Specify the range where you want to apply the conditional formatting (C3:Z8 based on the provided data)
start_row, end_row = worksheet.max_row - int(myData.get("maxTokenToSubscribe")), worksheet.max_row
start_col, end_col = 3, worksheet.max_column

# Apply conditional formatting based on the formula
for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        current_cell = worksheet.cell(row=row, column=col)
        previous_cell = worksheet.cell(row=row, column=col - 1)

        # Check if both cells contain numeric values
        if is_numeric(previous_cell.value) and is_numeric(current_cell.value):
            # Check if previous column value is less than current column value
            if float(previous_cell.value) < float(current_cell.value):
                current_cell.fill = green_fill
            elif float(previous_cell.value) >= float(current_cell.value):
                current_cell.fill = red_fill
        # elif current_cell.value == 0:
        #     # If current column value is zero, copy the fill color of the previous column
        #     current_cell.fill = green_fill

workbook.save('BNData.xlsx')
print("QUOTE")
# for ii in range(50):
#     t = datetime.datetime.now()
#     print(t.second)
#     print(t.microsecond/1000)
#     print(kite.quote("NSE:NIFTY BANK")["NSE:NIFTY BANK"]['last_price'])

# # Get instruments
# with open('readme.csv', 'w') as f:
#     f.write(str(kite.instruments('NFO')))
#
